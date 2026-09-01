#!/usr/bin/env python3
"""Erzeugt die Excel-Vorlage für XRechnung Batch Pro (Vorlage-Rechnungen.xlsx)."""
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = sys.argv[1] if len(sys.argv) > 1 else "Vorlage-Rechnungen.xlsx"

ACCENT = "1D5FBF"
LIGHT = "EEF3FB"
wb = Workbook()

head_font = Font(bold=True, color="FFFFFF", size=11)
head_fill = PatternFill("solid", fgColor=ACCENT)
sub_fill = PatternFill("solid", fgColor=LIGHT)
thin = Border(bottom=Side(style="thin", color="D0D7E2"))

# ---------- Blatt: Anleitung ----------
ws = wb.active
ws.title = "Anleitung"
ws.column_dimensions["A"].width = 100
rows = [
    ("XRechnung Batch Pro — Excel-Vorlage", 14, True),
    ("", 11, False),
    ("So funktioniert es:", 11, True),
    ("1. Blatt „Absender“: Ihre Firmendaten einmalig eintragen (gelten für alle Rechnungen).", 11, False),
    ("2. Blatt „Rechnungen“: pro Zeile eine Rechnungsposition. Mehrere Positionen einer Rechnung = mehrere Zeilen mit derselben Rechnungsnummer (Rechnungs- und Empfängerfelder nur in der ersten Zeile nötig).", 11, False),
    ("3. Datei speichern und in XRechnung-Batch-Pro.html ziehen → alle XRechnungen als ZIP.", 11, False),
    ("", 11, False),
    ("Hinweise:", 11, True),
    ("• Datum als Excel-Datum oder TT.MM.JJJJ. Beträge sind Netto-Einzelpreise; Dezimalkomma ist ok.", 11, False),
    ("• Einheiten: Stück, Stunde, Tag, Monat, pauschal, kg, m, m², Liter, km … (oder UN/ECE-Code).", 11, False),
    ("• Leitweg-ID/Referenz ist Pflicht: bei Behörden die Leitweg-ID, sonst z. B. Kundennummer.", 11, False),
    ("• Kleinunternehmer (§ 19 UStG): im Blatt „Absender“ auf „ja“ setzen — USt-Spalte wird dann ignoriert.", 11, False),
    ("• Die zwei Beispielrechnungen im Blatt „Rechnungen“ einfach überschreiben oder löschen.", 11, False),
]
for i, (text, size, bold) in enumerate(rows, start=1):
    c = ws.cell(row=i, column=1, value=text)
    c.font = Font(bold=bold, size=size, color=ACCENT if size == 14 else "1A2332")
    c.alignment = Alignment(wrap_text=True, vertical="top")

# ---------- Blatt: Absender ----------
ws = wb.create_sheet("Absender")
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 46
ws.append(["Feld", "Wert"])
for c in ws[1]:
    c.font = head_font; c.fill = head_fill
absender = [
    ("Name/Firma", "Muster Webdesign"),
    ("Ansprechpartner", "Max Muster"),
    ("Straße & Hausnummer", "Beispielweg 12"),
    ("PLZ", "67663"),
    ("Ort", "Kaiserslautern"),
    ("E-Mail", "rechnung@example.de"),
    ("Telefon", "+49 631 1234567"),
    ("USt-IdNr.", "DE123456789"),
    ("Steuernummer (falls keine USt-IdNr.)", ""),
    ("IBAN", "DE02120300000000202051"),
    ("BIC", ""),
    ("Kontoinhaber", "Max Muster"),
    ("Kleinunternehmer (ja/nein)", "nein"),
    ("Hinweistext Steuerbefreiung", "Gemäß § 19 UStG wird keine Umsatzsteuer berechnet."),
    ("Zahlungsbedingungen (Standard)", "Zahlbar innerhalb von 14 Tagen ohne Abzug."),
]
for k, v in absender:
    ws.append([k, v])
for row in ws.iter_rows(min_row=2):
    row[0].font = Font(bold=True)
    for c in row:
        c.border = thin
        c.alignment = Alignment(vertical="center")
ws.freeze_panes = "A2"

# ---------- Blatt: Rechnungen ----------
ws = wb.create_sheet("Rechnungen")
headers = ["Rechnungsnummer", "Rechnungsdatum", "Fällig am", "Zahlungsbedingungen",
           "Leitweg-ID / Referenz", "Empfänger Name", "Empfänger Straße", "Empfänger PLZ",
           "Empfänger Ort", "Empfänger E-Mail", "Bezeichnung", "Beschreibung",
           "Menge", "Einheit", "Einzelpreis netto", "USt %"]
widths = [18, 15, 12, 30, 22, 22, 22, 12, 16, 24, 28, 26, 9, 11, 15, 8]
ws.append(headers)
for i, (c, w) in enumerate(zip(ws[1], widths), start=1):
    c.font = head_font; c.fill = head_fill
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.column_dimensions[get_column_letter(i)].width = w
example = [
    ["RE-2026-001", "01.09.2026", "15.09.2026", "", "KD-1001", "Beispiel GmbH", "Industriestraße 5", "80331", "München", "buchhaltung@beispiel.de", "Webdesign-Leistungen", "Relaunch Startseite", 12, "Stunde", 95, 19],
    ["", "", "", "", "", "", "", "", "", "", "Hosting-Paket Business", "", 1, "Monat", 29.9, 19],
    ["RE-2026-002", "01.09.2026", "", "Zahlbar innerhalb von 7 Tagen.", "Bestellung 4711", "Handwerk Meier e.K.", "Werkstattweg 2", "55116", "Mainz", "", "Beratung vor Ort", "", 3.5, "Stunde", 110, 19],
]
for r in example:
    ws.append(r)
for row in ws.iter_rows(min_row=2, max_row=4):
    for c in row:
        c.border = thin
        if row[0].value:
            c.fill = sub_fill
ws.freeze_panes = "A2"
for col, fmt in ((13, "0.##"), (15, '#,##0.00" €"'), (16, "0")):
    for r in range(2, 400):
        ws.cell(row=r, column=col).number_format = fmt

wb.save(OUT)
print(f"Vorlage geschrieben: {OUT}")
